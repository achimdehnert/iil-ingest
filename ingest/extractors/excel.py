"""Excel extractor using openpyxl."""
from __future__ import annotations

from ingest.types import ExtractedContent

SUPPORTED_MIMES: frozenset[str] = frozenset({
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
})


class ExcelExtractor:
    """Extract rows and text from Excel bytes (requires openpyxl)."""

    supported_mimes = SUPPORTED_MIMES

    def can_handle(self, mime_type: str) -> bool:
        return mime_type in self.supported_mimes

    def extract(self, data: bytes) -> ExtractedContent:
        try:
            import openpyxl  # type: ignore[import]
        except ImportError as exc:  # pragma: no cover
            raise ImportError(
                "openpyxl required. Install with: pip install iil-ingest[excel]"
            ) from exc

        import io

        errors: list[str] = []
        text_parts: list[str] = []
        tables: list[list[list[str]]] = []
        sheet_names: list[str] = []

        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            for name in sheet_names:
                ws = wb[name]
                rows: list[list[str]] = []
                for row in ws.iter_rows(values_only=True):
                    str_row = [str(c) if c is not None else "" for c in row]
                    rows.append(str_row)
                    text_parts.append(" ".join(str_row))
                tables.append(rows)
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Excel read error: {exc}")

        return ExtractedContent(
            raw_bytes=data,
            text="\n".join(text_parts),
            tables=tables,
            metadata={"sheet_names": sheet_names},
            mime_type=next(iter(SUPPORTED_MIMES)),
            page_count=len(tables),
            extraction_errors=errors,
        )
